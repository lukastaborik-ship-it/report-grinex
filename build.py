#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Grinex LinkedIn dashboard — datová vrstva.

Najde NEJNOVĚJŠÍ soubor 'Grinex ContentPlan*.xlsx' ve složce o úroveň výš
(nebo v aktuální), vyčistí data dle pevných pravidel a vygeneruje data.json,
který načítá index.html. Vzhled dashboardu se nemění — mění se jen data.json.

Spuštění:  python3 build.py
"""

import json, glob, os, datetime, sys
from collections import defaultdict, Counter

try:
    import openpyxl
except ImportError:
    sys.exit("Chybí knihovna openpyxl. Nainstaluj: pip3 install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
SEARCH_DIRS = [os.path.join(HERE, ".."), HERE]
TODAY = datetime.datetime.now().replace(hour=23, minute=59, second=59, microsecond=0)

WEEKDAYS = ["Pondělí", "Úterý", "Středa", "Čtvrtek", "Pátek", "Sobota", "Neděle"]
WEEKDAYS_SHORT = ["Po", "Út", "St", "Čt", "Pá", "So", "Ne"]
MONTHS = ["Leden", "Únor", "Březen", "Duben", "Květen", "Červen",
          "Červenec", "Srpen", "Září", "Říjen", "Listopad", "Prosinec"]
# Osoby v sekci sítě (Network grow)
PERSONS = ["Richard Jahoda", "Richard Jahoda ml.", "Kamila Blechová", "Lenka Nečasová", "Grinex LinkedIn"]
# Ambasadoři — pro content performance (filtr/grafy)
AMBASSADORS = ["Richard Jahoda", "Richard Jahoda ml.", "Kamila Blechová", "Lenka Nečasová"]
LOW_SAMPLE = 10          # práh pro upozornění na malý vzorek
TIMING_OUTLIER_CAP = 35_000  # příspěvky nad tento dosah se z výpočtu nejlepšího času vyloučí


# ----------------------------------------------------------------------------
def find_workbook():
    cands = []
    for d in SEARCH_DIRS:
        cands += glob.glob(os.path.join(d, "Grinex ContentPlan*.xlsx"))
    # vyřaď dočasné lock soubory (~$...)
    cands = [c for c in cands if not os.path.basename(c).startswith("~$")]
    if not cands:
        sys.exit("Nenalezen žádný soubor 'Grinex ContentPlan*.xlsx'.")
    cands.sort(key=lambda p: os.path.getmtime(p))
    return cands[-1]  # nejnovější dle data úpravy


def header_map(ws, header_row):
    m = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip():
            m.setdefault(v.strip(), c)
    return m


def find_header_row(ws, must_have, max_scan=12):
    """Najde řádek hlavičky podle výskytu klíčových názvů sloupců."""
    for r in range(1, max_scan + 1):
        vals = {str(ws.cell(r, c).value).strip() for c in range(1, ws.max_column + 1)
                if ws.cell(r, c).value is not None}
        if all(h in vals for h in must_have):
            return r
    return None


def num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


# ----------------------------------------------------------------------------
def parse_posts(wb):
    ws = wb["📃Content plan "]
    hr = find_header_row(ws, ["Name of profile", "Status", "N. IMPRESSIONS"])
    if hr is None:
        sys.exit("Nenalezena hlavička listu 'Content plan'.")
    col = header_map(ws, hr)

    def g(r, name):
        c = col.get(name)
        return ws.cell(r, c).value if c else None

    posts = []
    status_counter = Counter()
    total_prepared = 0
    for r in range(hr + 1, ws.max_row + 1):
        name = g(r, "Name of profile")
        if name in (None, "", 0) or not isinstance(name, str):
            continue
        total_prepared += 1
        status = g(r, "Status")
        status_counter[status or "—"] += 1
        if status != "1 - Done":
            continue
        # datum: přednostně plánované datum (správně pokrývá všechny měsíce),
        # "DATE OF ACTUALITY" jen jako záloha, pokud plánovaný datum chybí
        d = g(r, "date of posting - plan") or g(r, "DATE OF ACTUALITY")
        imp = num(g(r, "N. IMPRESSIONS"))
        if not isinstance(d, datetime.datetime) or imp is None or imp <= 0:
            continue
        if d > TODAY:           # budoucí příspěvky se do výkonu nepočítají
            continue
        posts.append({
            "name": name.strip(),
            "date": d,
            "imp": int(imp),
            "likes": int(num(g(r, "N. LIKES")) or 0),
            "comments": int(num(g(r, "N. COMENTS + repost")) or 0),
            "idea": (g(r, " Main Idea") or g(r, "Main Idea") or "").strip() if isinstance(g(r, " Main Idea") or g(r, "Main Idea"), str) else "",
        })
    return posts, status_counter, total_prepared


def parse_network(wb):
    ws = wb["📈 Network grow"]
    hr = find_header_row(ws, ["Name", "Platform", "Date", "Followers", "Connections"])
    if hr is None:
        sys.exit("Nenalezena hlavička listu 'Network grow'.")
    col = header_map(ws, hr)

    # mapování názvů z tabulky na výstupní názvy
    NAME_MAP = {"Richard Jahoda st.": "Richard Jahoda"}

    def g(r, name):
        c = col.get(name)
        return ws.cell(r, c).value if c else None

    rows = []
    for r in range(hr + 1, ws.max_row + 1):
        raw_name = g(r, "Name")
        if raw_name is None or not isinstance(raw_name, str):
            continue
        name = NAME_MAP.get(raw_name.strip(), raw_name.strip())
        if name not in PERSONS:        # vyřaď #REF!, prázdné, neznámé
            continue
        d = g(r, "Date")
        foll = num(g(r, "Followers"))
        conn = num(g(r, "Connections"))
        if not isinstance(d, datetime.datetime):
            continue
        if foll is None and conn is None:   # budoucí prázdné týdny
            continue
        if d > TODAY:
            continue
        rows.append({
            "name": name,
            "platform": (g(r, "Platform") or "").strip(),
            "date": d, "foll": foll, "conn": conn,
        })
    return rows


# ----------------------------------------------------------------------------
def build(posts, status_counter, total_prepared, net):
    years = sorted({p["date"].year for p in posts})

    # ---- KPI a časové řady pro (rok|all) × (osoba|all) ----
    def subset(year, person):
        return [p for p in posts
                if (year == "all" or p["date"].year == year)
                and (person == "all" or p["name"] == person)]

    persons_in_posts = ["all"] + [pp for pp in AMBASSADORS if any(p["name"] == pp for p in posts)]

    kpis = {}
    monthly = {}        # dosah po měsících
    monthly_posts = {}  # počty příspěvků po měsících
    monthly_stacked = {}  # dosah po měsících rozdělený na osoby
    for yr in (["all"] + years):
        for person in persons_in_posts:
            key = f"{yr}|{person}"
            sub = subset(yr, person)
            tot = sum(p["imp"] for p in sub)
            n = len(sub)
            likes = sum(p["likes"] for p in sub)
            comments = sum(p["comments"] for p in sub)
            kpis[key] = {
                "reach": tot, "posts": n,
                "avg": round(tot / n) if n else 0,
                "likes": likes, "comments": comments,
                "engagement": round((likes + comments) / tot * 100, 2) if tot else 0,
            }
            # měsíční řada (jen pro konkrétní rok, ne all)
            if yr != "all":
                m_reach = [0] * 12
                m_posts = [0] * 12
                for p in sub:
                    m_reach[p["date"].month - 1] += p["imp"]
                    m_posts[p["date"].month - 1] += 1
                monthly[key] = m_reach
                monthly_posts[key] = m_posts

    # stacked po měsících (rok × {osoba: [12]})
    for yr in years:
        stacks = {}
        for person in AMBASSADORS:
            arr = [0] * 12
            for p in posts:
                if p["date"].year == yr and p["name"] == person:
                    arr[p["date"].month - 1] += p["imp"]
            if any(arr):
                stacks[person] = arr
        monthly_stacked[str(yr)] = stacks

    # ---- roční souhrn dosahu (pro srovnání let) ----
    yearly = {}
    for yr in years:
        yearly[str(yr)] = {
            amb: sum(p["imp"] for p in posts if p["date"].year == yr and p["name"] == amb)
            for amb in AMBASSADORS
        }

    # ---- nejlepší den / hodina / měsíc (průměrný dosah) ----
    def timing(year, person):
        sub = subset(year, person)
        # Outlier příspěvky (imp > TIMING_OUTLIER_CAP) by zkreslily průměry →
        # do výpočtu časování vstupují jen příspěvky v rámci normálního dosahu.
        sub_timing = [p for p in sub if p["imp"] <= TIMING_OUTLIER_CAP]
        outliers_excluded = len(sub) - len(sub_timing)
        byday = defaultdict(list)
        byhour = defaultdict(list)
        bymonth = defaultdict(list)
        heat = defaultdict(list)  # (weekday,hour)
        for p in sub_timing:
            d = p["date"]
            byday[d.weekday()].append(p["imp"])
            bymonth[d.month - 1].append(p["imp"])
            # Posts with no time info in Excel arrive at 00:00:00 — skip for hour/heatmap
            has_time = d.hour != 0 or d.minute != 0 or d.second != 0
            if has_time:
                byhour[d.hour].append(p["imp"])
                heat[(d.weekday(), d.hour)].append(p["imp"])
        day = [{"label": WEEKDAYS[i], "short": WEEKDAYS_SHORT[i],
                "avg": round(sum(byday[i]) / len(byday[i])) if byday[i] else 0,
                "n": len(byday[i]), "low": len(byday[i]) < LOW_SAMPLE}
               for i in range(7)]
        hour = [{"hour": h,
                 "avg": round(sum(byhour[h]) / len(byhour[h])) if byhour[h] else 0,
                 "n": len(byhour[h]), "low": len(byhour[h]) < LOW_SAMPLE}
                for h in range(24)]
        month = [{"label": MONTHS[i],
                  "avg": round(sum(bymonth[i]) / len(bymonth[i])) if bymonth[i] else 0,
                  "n": len(bymonth[i]), "low": len(bymonth[i]) < LOW_SAMPLE}
                 for i in range(12)]
        heatmap = []
        for wd in range(7):
            for h in range(24):
                vals = heat[(wd, h)]
                if vals:
                    heatmap.append({"d": wd, "h": h,
                                    "avg": round(sum(vals) / len(vals)), "n": len(vals)})
        return {"day": day, "hour": hour, "month": month, "heatmap": heatmap,
                "outliers_excluded": outliers_excluded, "outlier_cap": TIMING_OUTLIER_CAP}

    timing_data = {}
    for yr in (["all"] + years):
        for person in persons_in_posts:
            timing_data[f"{yr}|{person}"] = timing(yr, person)

    # ---- kumulativní dosah pro zvolený rok ----
    cumulative = {}
    for yr in years:
        m_reach = monthly.get(f"{yr}|all", [0] * 12)
        run = 0
        cum = []
        for v in m_reach:
            run += v
            cum.append(run)
        cumulative[str(yr)] = {"monthly": m_reach, "cumulative": cum}

    # ---- TOP příspěvky ----
    top = sorted(posts, key=lambda p: -p["imp"])[:20]
    top_posts = [{
        "name": p["name"], "date": p["date"].strftime("%d.%m.%Y"),
        "idea": p["idea"][:90], "imp": p["imp"],
        "likes": p["likes"], "comments": p["comments"],
    } for p in top]

    # ---- SÍŤ: časové řady + souhrny ----
    network = {}
    for person in PERSONS:
        platforms = sorted({r["platform"] for r in net if r["name"] == person})
        per_plat = {}
        for plat in platforms:
            pts = sorted([r for r in net if r["name"] == person and r["platform"] == plat],
                         key=lambda r: r["date"])
            series = [{"date": r["date"].strftime("%Y-%m-%d"),
                       "foll": r["foll"], "conn": r["conn"]} for r in pts]
            # výchozí / aktuální + přírůstky po letech
            foll_pts = [r for r in pts if r["foll"] is not None]
            conn_pts = [r for r in pts if r["conn"] is not None]
            summary = {}
            if foll_pts:
                summary["foll_start"] = foll_pts[0]["foll"]
                summary["foll_now"] = foll_pts[-1]["foll"]
                summary["foll_gain"] = foll_pts[-1]["foll"] - foll_pts[0]["foll"]
                summary["foll_pct"] = round((foll_pts[-1]["foll"] - foll_pts[0]["foll"]) / foll_pts[0]["foll"] * 100) if foll_pts[0]["foll"] else 0
            if conn_pts:
                summary["conn_start"] = conn_pts[0]["conn"]
                summary["conn_now"] = conn_pts[-1]["conn"]
                summary["conn_gain"] = conn_pts[-1]["conn"] - conn_pts[0]["conn"]
                summary["conn_pct"] = round((conn_pts[-1]["conn"] - conn_pts[0]["conn"]) / conn_pts[0]["conn"] * 100) if conn_pts[0]["conn"] else 0
            # přírůstek followers po letech (rozdíl konec roku - předchozí konec)
            yearly_foll = {}
            prev = None
            yrs = sorted({r["date"].year for r in foll_pts})
            for y in yrs:
                yend = [r for r in foll_pts if r["date"].year == y][-1]["foll"]
                base = prev if prev is not None else foll_pts[0]["foll"]
                yearly_foll[str(y)] = yend - base
                prev = yend
            summary["foll_yearly_gain"] = yearly_foll
            per_plat[plat] = {"series": series, "summary": summary}
        network[person] = per_plat

    # ---- pipeline (statusy) ----
    pipeline = {str(k): v for k, v in status_counter.items()}

    return {
        "meta": {
            "years": years,
            "persons": persons_in_posts,
            "total_prepared": total_prepared,
            "total_published": sum(1 for p in posts),
            "total_reach_all": sum(p["imp"] for p in posts),
        },
        "kpis": kpis,
        "monthly": monthly,
        "monthly_posts": monthly_posts,
        "monthly_stacked": monthly_stacked,
        "cumulative": cumulative,
        "yearly": yearly,
        "timing": timing_data,
        "top_posts": top_posts,
        "network": network,
        "pipeline": pipeline,
    }


# ----------------------------------------------------------------------------
def main():
    path = find_workbook()
    print(f"Načítám: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    posts, status_counter, total_prepared = parse_posts(wb)
    net = parse_network(wb)

    data = build(posts, status_counter, total_prepared, net)
    data["meta"]["source_file"] = os.path.basename(path)
    data["meta"]["generated_at"] = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    data["meta"]["data_until"] = TODAY.strftime("%d.%m.%Y")

    # Merge ručně zadaná LinkedIn Analytics data (li_analytics.json)
    li_path = os.path.join(HERE, "li_analytics.json")
    if os.path.exists(li_path):
        with open(li_path, encoding="utf-8") as fh:
            data["linkedin_analytics"] = json.load(fh)
        print(f"  LinkedIn Analytics: načteno {len(data['linkedin_analytics'])} profil(ů) z li_analytics.json")

    # Merge Meta (Facebook + Instagram) analytická data (meta_analytics.json)
    meta_path = os.path.join(HERE, "meta_analytics.json")
    if os.path.exists(meta_path):
        with open(meta_path, encoding="utf-8") as fh:
            data["meta_analytics"] = json.load(fh)
        print(f"  Meta Analytics: načteno z meta_analytics.json")

    # Merge YouTube analytická data (youtube_analytics.json)
    yt_path = os.path.join(HERE, "youtube_analytics.json")
    if os.path.exists(yt_path):
        with open(yt_path, encoding="utf-8") as fh:
            data["youtube_analytics"] = json.load(fh)
        print(f"  YouTube Analytics: načteno z youtube_analytics.json")

    # Merge Podcast analytická data (podcast_analytics.json)
    pod_path = os.path.join(HERE, "podcast_analytics.json")
    if os.path.exists(pod_path):
        with open(pod_path, encoding="utf-8") as fh:
            data["podcast_analytics"] = json.load(fh)
        print(f"  Podcast Analytics: načteno z podcast_analytics.json")

    out = os.path.join(HERE, "data.json")
    with open(out, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=1)

    # kontrolní výpis pro ověření přesnosti
    m = data["meta"]
    print(f"  Připravených příspěvků: {m['total_prepared']}")
    print(f"  Publikovaných (do výkonu): {m['total_published']}")
    print(f"  Celkový dosah (publ.): {m['total_reach_all']:,}")
    print(f"  Roky: {m['years']}")
    for person, plats in data["network"].items():
        for plat, d in plats.items():
            s = d["summary"]
            print(f"  Síť {person}/{plat}: foll {s.get('foll_start')}→{s.get('foll_now')} | conn {s.get('conn_start')}→{s.get('conn_now')}")
    print(f"Hotovo → {out}")


if __name__ == "__main__":
    main()
